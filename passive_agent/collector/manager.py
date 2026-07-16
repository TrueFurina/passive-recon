"""采集管理器 — 编排多数据源并行采集、去重、聚合。

一通百通：任意目标（高校/企业/政府/域名）→ 自动推断域名 → 全源采集。
"""
from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from passive_agent.collector.domain_db import infer_domain, verify_domain_alive
from passive_agent.collector.model import AssetRecord, AssetSourceEnum, AssetType, CollectReport
from passive_agent.collector.sources import (
    CrtshCollector,
    FofaCollector,
    HackerTargetCollector,
    HunterCollector,
    OTXCollector,
    QichachaCollector,
    ReverseDnsCollector,
    SecurityTrailsCollector,
    URLScanCollector,
)
from passive_agent.common import logging as clog
from passive_agent.config import settings

_logger = clog.get_logger("collector-manager")

# 支持的被动数据源清单（名称 → 说明）
SUPPORTED_SOURCES = {
    "crt.sh": "证书透明日志（免费）",
    "hackertarget": "HackerTarget API（免费）",
    "otx": "AlienVault OTX 被动 DNS（免费/需Key）",
    "urlscan": "URLScan.io 历史快照（免费）",
    "securitytrails": "SecurityTrails 子域名（需Key）",
    "hunter": "鹰图空间测绘（需Key）",
    "fofa": "FOFA 空间搜索引擎（需email+Key）",
    "reverse_dns": "IP 反查 DNS（免费）",
    "qichacha": "企查查 企业工商信息（需Key）",
}


class CollectorManager:
    """被动资产采集管理器 — 一通百通。

    用法:
        # 自动推断域名
        mgr = CollectorManager()
        report = mgr.collect("北京大学")
        report = mgr.collect("阿里巴巴")

        # 手动指定域名
        report = mgr.collect("目标企业", domain="target.com")

        # 指定数据源
        report = mgr.collect("武汉大学", enabled_sources=["crt.sh", "hackertarget"])
    """

    def __init__(self) -> None:
        self._errors: List[str] = []

    # ── 采集入口（一通百通）──

    def collect(self, name: str, domain: str = "",
                enabled_sources: Optional[List[str]] = None,
                known_ips: Optional[List[str]] = None) -> CollectReport:
        """对任意目标执行全被动资产采集。

        一通百通：输入"北京大学"自动推断为 pku.edu.cn，全源采集。

        Args:
            name: 目标中文/英文名称（企业名/高校名/域名）
            domain: 主域名（可选，不传则自动推断）
            enabled_sources: 启用的数据源，None=全部
            known_ips: 已知 IP 列表（可选，用于反向 DNS）
        """
        # 自动推断域名
        if not domain:
            domain = infer_domain(name)
            _logger.info(f"自动推断域名: {name} → {domain}")

        report = CollectReport(enterprise=name, domain=domain)
        collectors = self._build_collectors(domain, enabled_sources, known_ips)

        _logger.info(f"开始采集: {name} ({domain})，启用 {len(collectors)} 个数据源")

        # 并发采集（ThreadPoolExecutor 并行多个数据源）
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading

        _results_lock = threading.Lock()

        def _run_one(source_name: str, collector) -> tuple:
            try:
                records = collector.collect(domain)
                return (source_name, records, None)
            except PermissionError as e:
                return (source_name, [], f"[{source_name}] R1 拦截: {e}")
            except Exception as e:
                return (source_name, [], f"[{source_name}] 异常: {e}")

        with ThreadPoolExecutor(max_workers=len(collectors)) as pool:
            fut_map = {pool.submit(_run_one, n, c): n for n, c in collectors}
            for fut in as_completed(fut_map):
                source_name, records, err = fut.result()
                with _results_lock:
                    for r in records:
                        report.records.append(r)
                    report.sources_used.append(source_name)
                if err:
                    with _results_lock:
                        self._errors.append(err)
                        report.errors.append(err)
                    _logger.warn(err)
                else:
                    _logger.info(f"  [{source_name}] → {len(records)} 条")

        # 去重 + IP 补全 + 风险检测
        self._dedup(report)
        self._enrich_ips(report)
        self._detect_risks(report)
        report.total_records = len(report.records)
        report.completed_at = datetime.now(timezone.utc).isoformat()

        return report

    # ── Excel 导出 ──

    def export_to_excel(self, report: CollectReport, path: str = "report.xlsx") -> str:
        """将采集报告导出为 Excel（多 sheet）。"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()

        # Sheet1: 资产清单
        ws1 = wb.active
        ws1.title = "资产清单"
        ws1.append(["资产值", "类型", "数据源", "IP", "端口", "技术栈", "标题", "标签"])
        for r in report.records:
            ws1.append([
                r.value, r.asset_type.value, r.source.value,
                r.ip or "", r.port or "",
                ", ".join(r.tech_stack) if r.tech_stack else "",
                r.title, ", ".join(r.tags),
            ])

        # Sheet2: 统计总览
        ws2 = wb.create_sheet("统计总览")
        ws2.append(["指标", "数值"])
        type_counts = {}
        ips = set()
        for r in report.records:
            type_counts[r.asset_type.value] = type_counts.get(r.asset_type.value, 0) + 1
            if r.ip: ips.add(r.ip)
        ws2.append(["目标", f"{report.enterprise} ({report.domain})"])
        ws2.append(["数据源", ", ".join(report.sources_used)])
        ws2.append(["总资产", report.total_records])
        ws2.append(["IP数", len(ips)])
        for t, c in sorted(type_counts.items(), key=lambda x: -x[1]):
            ws2.append([f"{t}数", c])
        ws2.append(["执行时间", report.completed_at])
        if report.errors:
            ws2.append(["风险/错误", len(report.errors)])

        # Sheet3: 风险发现
        if report.errors:
            ws3 = wb.create_sheet("风险发现")
            ws3.append(["风险描述"])
            for e in report.errors:
                ws3.append([e])

        wb.save(path)
        _logger.info(f"Excel 已导出: {path}")
        return path

    # ── 通用导入：从已有资产目录/文件导入种子数据 ──

    def import_from_dir(self, path: str, enterprise: str = "",
                        domain: str = "") -> CollectReport:
        """从已有资产目录导入种子数据（Markdown / JSON / CSV）。

        通用版，不限于 FAFU 格式。支持：
        - Markdown 表格（| domain | ip | 格式）
        - JSON 文件（{"资产规模": {...}, "高风险发现": [...]}）
        """
        base = os.path.abspath(path)
        if not os.path.isdir(base):
            return self._import_from_file(path, enterprise, domain)

        # 从目录名推断目标
        if not enterprise:
            enterprise = os.path.basename(base)
        if not domain:
            domain = infer_domain(enterprise)

        report = CollectReport(enterprise=enterprise, domain=domain)

        # 1. 读取所有 .md 文件，提取表格中的域名/IP
        seen: set = set()
        for fname in os.listdir(base):
            if not fname.endswith((".md", ".txt")):
                continue
            fpath = os.path.join(base, fname)
            try:
                with open(fpath, "r", encoding="utf-8") as fh:
                    content = fh.read()
                for line in content.split("\n"):
                    if "|" not in line:
                        continue
                    parts = [p.strip() for p in line.split("|")]
                    for p in parts:
                        if p.endswith(f".{domain.split('.', 1)[1]}") if "." in domain else False:
                            if p not in seen and '.' in p:
                                seen.add(p)
                                ip = ""
                                for p2 in parts:
                                    if p2.count(".") == 3 and all(c.isdigit() or c == "." for c in p2):
                                        ip = p2
                                        break
                                report.records.append(AssetRecord(
                                    value=p, asset_type=AssetType.SUBDOMAIN,
                                    source=AssetSourceEnum.CUSTOM, ip=ip,
                                    tags=["seed"],
                                ))
            except Exception as exc:
                _logger.warn(f"读取 {fname} 失败: {exc}")

        # 2. 读取 JSON 报告（通用格式）
        json_path = os.path.join(base, f"比赛报告_{enterprise}.json" if enterprise else "")
        if not os.path.exists(json_path):
            json_path = os.path.join(base, "report.json")
        if os.path.exists(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as fh:
                    data = json.load(fh)
                report.domain = data.get("主域名", report.domain)
                assets = data.get("资产规模", data.get("summary", {}))
                for key, val in assets.items() if isinstance(assets, dict) else []:
                    report.records.append(AssetRecord(
                        value=f"{key}: {val}",
                        asset_type=AssetType.UNKNOWN,
                        source=AssetSourceEnum.CUSTOM,
                        tags=["summary", key],
                    ))
                risks = data.get("高风险发现", data.get("risks", []))
                for r in risks:
                    r_level = r.get("等级", r.get("level", "?"))
                    r_desc = r.get("风险", r.get("description", str(r)))
                    r_impact = r.get("影响", r.get("impact", ""))
                    report.records.append(AssetRecord(
                        value=f"[{r_level}] {r_desc}",
                        asset_type=AssetType.UNKNOWN,
                        source=AssetSourceEnum.CUSTOM,
                        source_extra=r_impact,
                        tags=["risk", str(r_level).lower()],
                    ))
            except Exception as e:
                _logger.warn(f"读取 JSON 报告失败: {e}")

        _logger.info(f"通用导入: {path} → {len(seen)} 子域名")
        report.total_records = len(report.records)
        report.sources_used.append(AssetSourceEnum.CUSTOM.value)
        report.completed_at = datetime.now(timezone.utc).isoformat()
        return report

    def _import_from_file(self, file_path: str, enterprise: str = "",
                          domain: str = "") -> CollectReport:
        """从单个文件导入资产。"""
        if not enterprise:
            enterprise = os.path.splitext(os.path.basename(file_path))[0]
        if not domain:
            domain = infer_domain(enterprise)
        report = CollectReport(enterprise=enterprise, domain=domain)
        try:
            with open(file_path, "r", encoding="utf-8") as fh:
                content = fh.read()
            # 尝试解析 JSON
            try:
                data = json.loads(content)
                if isinstance(data, dict):
                    for k, v in data.items():
                        report.records.append(AssetRecord(
                            value=str(v), asset_type=AssetType.UNKNOWN,
                            source=AssetSourceEnum.CUSTOM,
                            tags=[k],
                        ))
            except json.JSONDecodeError:
                # 当作纯文本，提取域名
                import re
                domains = re.findall(r'[a-zA-Z0-9]([a-zA-Z0-9-]*[a-zA-Z0-9])?\.([a-zA-Z]{2,})', content)
                for d in set(domains):
                    report.records.append(AssetRecord(
                        value=d[0] + '.' + d[1],
                        asset_type=AssetType.DOMAIN,
                        source=AssetSourceEnum.CUSTOM,
                    ))
        except Exception as e:
            report.errors.append(f"文件导入失败: {e}")
        report.total_records = len(report.records)
        return report

    # ── 内部方法 ──

    def _build_collectors(self, domain: str,
                          enabled: Optional[List[str]],
                          known_ips: Optional[List[str]]) -> List:
        """构建采集器列表（按配置决定启用哪些）。"""
        api_keys = getattr(settings, "API_KEYS", {})

        all_collectors = [
            ("crt.sh", CrtshCollector(timeout=15)),
            ("hackertarget", HackerTargetCollector(timeout=15)),
            ("otx", OTXCollector(timeout=15, api_key=api_keys.get("otx", ""))),
            ("urlscan", URLScanCollector(timeout=15)),
            ("securitytrails", SecurityTrailsCollector(
                timeout=15, api_key=api_keys.get("securitytrails", ""))),
            ("hunter", HunterCollector(
                timeout=20, api_key=api_keys.get("hunter", ""))),
            ("fofa", FofaCollector(
                timeout=20, api_key=api_keys.get("fofa", ""))),
            ("qichacha", QichachaCollector(
                timeout=15, api_key=api_keys.get("qichacha", {}))),
        ]
        if known_ips:
            all_collectors.append(
                ("reverse_dns", ReverseDnsCollector(target_ips=known_ips, timeout=15)),
            )

        if enabled:
            return [(n, c) for n, c in all_collectors if n in enabled]
        return all_collectors

    @staticmethod
    def _dedup(report: CollectReport) -> None:
        """按 value + asset_type 去重，保留第一个。"""
        seen: set = set()
        unique: List[AssetRecord] = []
        for r in report.records:
            key = (r.value, r.asset_type.value)
            if key not in seen:
                unique.append(r)
                seen.add(key)
        report.records = unique

    @staticmethod
    def _enrich_ips(report: CollectReport) -> None:
        """对没有 IP 的子域名并发补 DNS 解析。"""
        import socket
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _resolve(d: str) -> tuple:
            try:
                addrs = socket.getaddrinfo(d, 80, socket.AF_INET, socket.SOCK_STREAM)
                ips = list(set(a[4][0] for a in addrs))
                return (d, ips[0] if ips else None)
            except Exception:
                return (d, None)

        targets = [(i, r) for i, r in enumerate(report.records)
                   if r.asset_type in (AssetType.SUBDOMAIN, AssetType.DOMAIN) and not r.ip]
        if not targets:
            return

        with ThreadPoolExecutor(max_workers=20) as pool:
            fut_map = {pool.submit(_resolve, r.value): idx for idx, r in targets}
            enriched = 0
            for fut in as_completed(fut_map, timeout=30):
                d, ip = fut.result()
                if ip:
                    idx = fut_map[fut]
                    report.records[idx].ip = ip
                    enriched += 1

        _logger.info(f"DNS 并发补全: {enriched}/{len(targets)} 条补充了 IP")

    @staticmethod
    def _detect_risks(report: CollectReport) -> None:
        """精准风险检测 — 只报真实高风险，不泛泛警告。"""
        risks = []

        # 1. 高危端口（真正有风险的）
        risk_ports = {
            21: "[P2] FTP 明文传输（可嗅探）",
            23: "[P1] Telnet 明文登录（可嗅探凭据）",
            3389: "[P2] RDP 远程桌面公网暴露",
            6379: "[P1] Redis 未授权风险（默认无密码）",
            27017: "[P1] MongoDB 公网暴露",
            5432: "[P2] PostgreSQL 公网暴露",
            3306: "[P2] MySQL 公网暴露",
            8080: "[P2] 管理后台/API 接口（非标准端口）",
            8443: "[P2] 管理后台 HTTPS（非标准端口）",
            9090: "[P2] 管理控制台（非标准端口）",
        }
        seen_ports = set()
        for r in report.records:
            if r.port and r.port in risk_ports and r.port not in seen_ports:
                seen_ports.add(r.port)
                risks.append(f"{risk_ports[r.port]}: {r.value}")

        # 2. 核心技术栈风险（只报有价值的信息）
        value_tech_map = {
            "CAS": "[P2] 统一认证系统: {}",
            "Coremail": "[P2] Coremail 邮件系统（默认密码策略风险）: {}",
            "Synjones": "[P0] Synjones 一卡通系统（已知设备默认密码）: {}",
            "Apache": "[P2] Web 服务: {}",
            "nginx": "",  # nginx 常见不报风险
            "IIS": "[P2] IIS Web 服务: {}",
            "Spring": "[P2] Spring Boot 应用: {}",
            "Joomla": "[P2] Joomla CMS（已知历史漏洞）: {}",
            "WordPress": "[P1] WordPress（常见插件漏洞）: {}",
            "DES": "[P1] 使用 DES 加密（已过时可破解）: {}",
            "通达OA": "[P1] 通达OA（已知历史漏洞）: {}",
            "致远OA": "[P1] 致远OA（已知历史漏洞）: {}",
        }
        seen_value = set()
        for r in report.records:
            for tech in r.tech_stack:
                tech_s = tech if isinstance(tech, str) else str(tech)
                for key, template in value_tech_map.items():
                    if key.lower() in tech_s.lower() and template:
                        dedup_key = f"{key}:{r.value}"
                        if dedup_key not in seen_value:
                            seen_value.add(dedup_key)
                            risks.append(template.format(r.value))

        # 3. 子域名分析 — 仅提真正的暴露面
        exposure_map = {
            "webvpn": "[P1] WebVPN 远程接入暴露",
            "vpn": "[P1] VPN 入口暴露",
            "oa": "[P2] OA 系统暴露",
            "mail": "[P2] 邮件系统暴露",
            "ecard": "[P1] 一卡通系统暴露",
            "erp": "[P2] ERP 系统暴露",
            "sso": "[P2] SSO 单点登录暴露",
            "cas": "[P2] CAS 统一认证暴露",
            "coremail": "[P2] Coremail 邮件系统暴露",
            "jira": "[P2] Jira 项目管理暴露",
            "jenkins": "[P1] Jenkins（已知未授权RCE风险）",
            "gitlab": "[P2] GitLab 源码管理暴露",
            "phpmyadmin": "[P1] phpMyAdmin 暴露（暴力破解风险）",
        }
        seen_exposure = set()
        for r in report.records:
            val_lower = r.value.lower()
            for keyword, warning in exposure_map.items():
                if keyword in val_lower:
                    dedup_key = keyword
                    if dedup_key not in seen_exposure:
                        seen_exposure.add(dedup_key)
                        risks.append(f"{warning}: {r.value}")

        # 最终去重输出
        for risk in sorted(set(risks)):
            report.errors.append(f"🔴 {risk}")
        _logger.info(f"精准风险检测: 发现 {len(set(risks))} 项")
